import numbers
import re
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import colors
from openpyxl import Workbook
from openpyxl.worksheet import pagebreak
import openpyxl.worksheet.header_footer
import datetime
from openpyxl.styles.numbers import builtin_format_code, builtin_format_id
from openpyxl.styles import numbers

import re

def merge_and_release(directory_set,save_dir,PO_table):

    wb = Workbook()
    ws=wb.active
    ws.title="요약 시트"

    for directory in directory_set:


        font1=Font(name='맑은 고딕',
                   size=18,
                   bold=True
                   )
        alignment=Alignment(horizontal="center",vertical="bottom",wrap_text=True)
        #추후 worksheet_name을 번호로 변경하도록



        # directory = "C:/Users/HAN/Documents/카카오톡 받은 파일/PO_4467200.xlsx"

        ws=wb.create_sheet(title=openpyxl.open(directory).sheetnames[0])

        '''셀 병합 및 가운데 맞춤 과정'''
        # ws.merge_cells('A1:E1')
        # ws.merge_cells('A2:B2')
        # ws.merge_cells('A3:B3')
        # ws.merge_cells('C2:E2')
        # ws.merge_cells('C3:E2')
        # ws.merge_cells('C4:E4')
        ''''''

        top_left_cell=ws['A1']

        top_left_cell.value="주식회사듀벨 (A00059636)"

        yellow_fill=PatternFill(start_color='E5D85C',end_color='E5D85C',fill_type='solid')
        def merge_and_centered(cells):
            ws[cells[0]].alignment=alignment
            ws[cells[0]].font=font1
            ws.merge_cells('{}:{}'.format(cells[0],cells[1]))

            if cells !=('A1','E1'):
                ws[cells[0]].fill=yellow_fill
        for cells in [('A1','E1'),('A2','B2'),('A3','B3'),('C2','E2'),('C3','E3'),('A4','E4')]:merge_and_centered(cells)
        ws['A1']="주식회사듀벨 (A00059636)"; ws["A2"]='입고예정일'; ws["A3"]='물류센터'; ws['A5']='순번'; ws['B5']='바코드'; ws['C5']='상품명'; ws['D5']='수량'; ws['E5']='비고'  #



        #행렬 너비지정
        for row in range(6, 21):
            ws.row_dimensions[row].height=40.5
        for row in range(1, 6):
            ws.row_dimensions[row].height=30
        ws.column_dimensions['A'].width=4.13
        ws.column_dimensions['B'].width=14
        ws.column_dimensions['C'].width=51.38
        ws.column_dimensions['D'].width=8.63
        ws.column_dimensions['E'].width=8.63

        #테두리지정
        def set_border(ws, cell_range):
            thin = Side(border_style="thin", color="000000")
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

        set_border(ws,"A1:E5")

        #데이터가지고오기

        df=pd.read_excel(directory)
        date_=datetime.datetime.strptime(' '.join(df.iloc[11][5].split()[0].split('/')),"%Y %m %d")
        ws['C2']=(date_.strftime("%Y년 %m월 %d일 {}".format(["월요일","화요일","수요일","목요일","금요일","토요일","일요일"][date_.weekday()])))
        if PO_table.loc[lambda x: x[('발주', '번호')] == df.iloc[8][2], ].shape[0]==1:
            transport=PO_table.loc[lambda x: x[('발주', '번호')] == df.iloc[8][2], ('발주','운송')].values[0]
        else:
            transport='사이트 내 확인 필요'

        ws['C3']=df.iloc[11][2]+f" - {transport}"

        num=df.index[df.iloc[:,0]=="No."][0]+2

        text_wraping=Alignment(wrap_text=True)
        total_quantity=0
        while df.iloc[num,0]!="합계":
            try :
                index=int(df.iloc[num][0])
                a=ws.cell(row=5+index,column=1,value=index)
                b=ws.cell(row=5+index,column=3,value=df.iloc[num][2])
                c=ws.cell(row=5+index,column=4,value=df.iloc[num][6])
                ws.cell(row=5+index,column=20,value=df.iloc[num][9])
                a.alignment=text_wraping
                b.alignment=text_wraping
                c.alignment=text_wraping
                c.font=Font(size=15)
                total_quantity+=int(c.value)

            except ValueError:
                ws.cell(row=5+index,column=2,value=df.iloc[num][2])

            num+=1
        ws['A4']="발주번호 : "+str(df.iloc[8][2])+"(합계수량 : {}개)".format(total_quantity)
        set_border(ws,"A1:E{}".format(5+index))

        #page 설정

        ws.page_setup.fitToPage=True
        ws.print_options.horizontalCentered=True
        ws.page_setup.paperSize=9
        ws.print_area="A1:E{}".format(5+index)
        ws.oddFooter.center.text="&P / &N"
        ws.oddHeader.center.text="&[Date]"

    #summery
    ws=wb['요약 시트']
    ws['A1']="요약 - {}".format(datetime.date.today())
    ws['A1'].font=font1
    ws['A1'].alignment=alignment
    ws['A1'].font=Font(size=20)
    ws.merge_cells('A1:I1')
    for i,j in enumerate(["순번","발주번호","바코드","상품명","물류센터","입고예정일","수량","매입가액","매입가액 X 수량"]):
        ws.cell(row=2,column=(1+i),value=j)

    for sheet in wb.sheetnames[1:]:
        ws1=wb[sheet]
        for i in ws1.iter_rows(min_row=6,max_col=21):
            current_row=ws.max_row+1
            order_number=ws1.cell(row=4,column=1).value[7:15]
            ship_center=ws1.cell(row=3,column=3).value[0:4]
            expected_stocked_day=ws1.cell(row=2,column=3).value

            ws.cell(row=current_row,column=1,value=current_row-2)
            ws.cell(current_row, 2, int(order_number)).number_format='0'
            ws.cell(current_row,3,int(i[1].value)).number_format=numbers.FORMAT_NUMBER
            ws.cell(current_row, 4,i[2].value).alignment=alignment
            ws.cell(current_row, 5, ship_center)
            ws.cell(current_row,6,expected_stocked_day).number_format=builtin_format_code(14)
            ws.cell(current_row,7,int(i[3].value)).number_format
            ws.cell(current_row,8,int(i[19].value))
            ws.cell(current_row,9,ws.cell(current_row,7).value * ws.cell(current_row,8).value)
            # ws.cell(row=5 + index, column=20, value=df.iloc[num][9])

            ws.row_dimensions[current_row].height=34
            ws.column_dimensions['A'].width=4.13
            ws.column_dimensions['B'].width=14
            ws.column_dimensions['C'].width=14
            ws.column_dimensions['D'].width=51.38
            ws.column_dimensions['E'].width=8
            ws.column_dimensions['F'].width=24
            ws.column_dimensions['G'].width=7
            ws.column_dimensions['H'].width=8
            ws.column_dimensions['I'].width=15


    set_border(ws,"A1:I{}".format(current_row))
    ws.row_dimensions[1].height=42

    #프린트설정
    ws.print_area="A1:I{}".format(current_row)
    ws.page_setup.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_setup.paperSize = 9

    ws.page_margins.left=0.26
    ws.page_margins.right=0.26

    #순서설정
    PO_table=pd.DataFrame
    PO_table.sort_values(by=('발주','번호'),ascending=False)

    wb.sheetnames
    wb.move_sheet()
    wb.save(save_dir)
